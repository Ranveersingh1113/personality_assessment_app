
import os
import sys
import json
import pandas as pd
from openpyxl import Workbook

# Ensure we can import from ai_core
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from ai_core.personality_assessment import PersonalityAssessmentSystem
from ai_core.report_card_generator import ReportCardGenerator

def create_mock_template():
    """Create a dummy template for testing"""
    print("Creating mock template...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Card"
    
    # Add dummy headers
    ws['B2'] = "नाव :"
    ws['J2'] = "शाळा :"
    
    # Ensure columns exist by writing to far right
    ws['AO1'] = "Border"
    
    ws['B4'] = "क्षमता (Strengths)"
    ws['G4'] = "कमतरता (Weaknesses)"
    ws['L4'] = "संधी (Opportunities)"
    ws['Q4'] = "भीती (Threats)"
    
    wb.save("report_card_template.xlsx")
    print("Mock template created.")

def verify_marathi_swot():
    print("\n--- Verifying Marathi SWOT ---")
    try:
        system = PersonalityAssessmentSystem()
        system.setup_vector_database()
        
        observations = "Student is very active, helps others but talks too much in class. Good at drawing."
        print(f"\nObservations: {observations}")
        
        print("Generating Marathi SWOT...")
        result = system.generate_marathi_swot(observations)
        print("Result:")
        print(json.dumps(result, indent=2, ensure_ascii=False))
        
        if result.get('swot_items'):
            print("✅ Marathi SWOT generated successfully.")
            return result
        else:
            print("❌ Marathi SWOT generation failed.")
            return None
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Error: {e}")
        return None

def verify_report_generation(swot_data):
    print("\n--- Verifying Report Card Generation ---")
    if not swot_data:
        print("Skipping report verification due to missing SWOT data.")
        return

    try:
        create_mock_template()
        
        gen = ReportCardGenerator()
        # No need to set template via upload, we verified file exists
        
        print("Generating single report...")
        path = gen.generate_report_card("Test Student", "Test School", swot_data)
        print(f"Report generated at: {path}")
        
        if os.path.exists(path):
            print("✅ Report file created.")
        else:
            print("❌ Report file missing.")
            
        print("Generating batch report...")
        batch_data = [{'name': 'Batch Student 1', 'school': 'School A', 'swot_data': swot_data}]
        zip_path, errors = gen.batch_generate(batch_data)
        
        if os.path.exists(zip_path):
            print(f"✅ Batch ZIP created locally at: {zip_path}")
        else:
            print("❌ Batch ZIP missing.")
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Error during report generation: {e}")

if __name__ == "__main__":
    swot_result = verify_marathi_swot()
    if swot_result:
        verify_report_generation(swot_result)
